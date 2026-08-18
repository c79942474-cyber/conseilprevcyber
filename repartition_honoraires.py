# -*- coding: utf-8 -*-
"""Le tableau de répartition des honoraires de maîtrise d'œuvre, en classeur.

CE QUE CE MODULE PRODUIT. Le tableau que la maîtrise d'ouvrage attend en pièce
de marché : phases MOP en lignes, cotraitants en colonnes, ventilation et
montant pour chacun. Il est rempli depuis `moe_dc`, jamais à la main.

═══════════════════════════════════════════════════════════════════════════
POURQUOI LE MODÈLE FOURNI N'EST PAS REPRIS TEL QUEL
═══════════════════════════════════════════════════════════════════════════

Le classeur d'origine porte ses formules, et l'idée naturelle était de ne
remplir que ses cases jaunes en laissant Excel calculer. Vérification faite
cellule par cellule, ces formules rendraient un total FAUX quoi qu'on y mette :

  — `D26` (« sous-total tranche optionnelle 2 ») vaut `=SUM(D19:D28)`. La
    plage CONTIENT D26 : c'est une référence circulaire. Elle englobe en outre
    d'autres sous-totaux (D20) et des lignes d'autres tranches (DCE, DAACT) ;
  — `D24` (« sous-total tranche optionnelle 1 ») n'a AUCUNE formule : le
    sous-total en montant n'existe pas, alors que `D38` l'additionne ;
  — `D35` vaut `=SUM(D30:D34)` — VISA à GPA — alors que sa tranche commence à
    DCE : elle omet DCE, DAACT et ACT ;
  — `C38` vaut `=C37+C35+C26+C24` et n'inclut PAS `C19` : la ventilation de
    l'esquisse n'entre jamais dans le total, qui ne peut donc pas faire 100 %.

À quoi s'ajoutent deux irrégularités : `D20` multiplie par `D12` (le taux
OPTIONNEL) au lieu de `D13` (la base), et `D19` porte 15 000 en dur quand
toutes les autres lignes sont des formules.

CE QU'ON EN FAIT. On garde la MISE EN PAGE et le VOCABULAIRE du modèle — c'est
ce que le lecteur reconnaît — et on écrit des VALEURS calculées par `moe_dc`,
vérifiables ligne à ligne. Les défauts relevés sont écrits dans le classeur
lui-même : les taire reviendrait à laisser quelqu'un remplir de nouveau
l'original en croyant qu'il calcule juste.

═══════════════════════════════════════════════════════════════════════════
CE QUE LE BARÈME NE SAIT PAS DÉCOUPER — ET QU'ON N'INVENTE PAS
═══════════════════════════════════════════════════════════════════════════

Le modèle demande quatorze lignes MOP. Le barème n'en connaît que cinq, parce
que le relevé dont il vient les groupe ainsi. Trois lignes du modèle sont donc
DES REGROUPEMENTS, et le partage à l'intérieur de chacun n'existe nulle part :

    ESQ + APS + DPC     ←  phase « aps » du barème
    PRO + DCE           ←  phase « pro »
    VISA + DET + AOR + DOE  ←  phase « exe »

Le montant est porté par la PREMIÈRE ligne du groupe, et les suivantes portent
« compris dans … ». Écrire 0 laisserait croire que rien n'est dû pour cette
phase ; répartir au prorata inventerait un partage que personne n'a publié.

Deux lignes du modèle n'ont AUCUN équivalent au barème — DAACT (conformité au
permis) et GPA (garantie de parfait achèvement). Elles restent vides et le
disent. C'est un manque du barème, pas du classeur.
"""
import io
import os
from datetime import date as _date

VERSION = "2026-08-a"

# Les lignes du modèle, dans son ordre et avec ses intitulés. `phase` nomme la
# phase du barème qui la couvre ; `groupe` dit qu'elle est comprise dans une
# ligne précédente ; `absent` qu'aucune phase du barème ne la couvre.
LIGNES = [
    {"code": "ESQ",   "nom": "Études d'esquisse",                          "phase": "aps"},
    {"code": "APS",   "nom": "Avant-projet sommaire",                      "groupe": "ESQ"},
    {"code": "DPC",   "nom": "Dossier de demande de permis de construire", "groupe": "ESQ"},
    {"code": "APD",   "nom": "Avant-projet définitif",                     "phase": "apd"},
    {"code": "PRO",   "nom": "Études de projet",                           "phase": "pro"},
    {"code": "DCE",   "nom": "Dossier de consultation des entreprises",    "groupe": "PRO"},
    {"code": "DAACT", "nom": "Conformité au permis de construire",         "absent": True},
    {"code": "ACT",   "nom": "Assistance à la passation des marchés de travaux",
     "phase": "act"},
    {"code": "VISA",  "nom": "Visa des études réalisées par l'entreprise", "phase": "exe"},
    {"code": "DET",   "nom": "Direction de l'exécution des contrats de travaux",
     "groupe": "VISA"},
    {"code": "AOR",   "nom": "Assistance aux opérations de réception",     "groupe": "VISA"},
    {"code": "DOE",   "nom": "Dossier des ouvrages exécutés",              "groupe": "VISA"},
    {"code": "GPA",   "nom": "Garantie de parfait achèvement",             "absent": True},
    # L'ordonnancement-pilotage-coordination est une MISSION au barème, pas une
    # phase : sa ligne se remplit depuis la mission du même nom, si elle est
    # retenue. Le modèle en fait une ligne de phase — on suit le modèle.
    {"code": "OPC",   "nom": "Ordonnancement, pilotage et coordination",   "mission": "opc"},
]

# Ce que chaque regroupement recouvre, pour l'écrire au lieu de le supposer.
GROUPES = {
    "ESQ":  ["ESQ", "APS", "DPC"],
    "PRO":  ["PRO", "DCE"],
    "VISA": ["VISA", "DET", "AOR", "DOE"],
}

RESERVES_MODELE = [
    "D26 « sous-total tranche optionnelle 2 » = SUM(D19:D28) : la plage "
    "contient D26 elle-même — référence circulaire — et englobe d'autres "
    "sous-totaux ainsi que des lignes d'autres tranches.",
    "D24 « sous-total tranche optionnelle 1 » ne porte aucune formule, alors "
    "que le total général D38 l'additionne.",
    "D35 = SUM(D30:D34) couvre VISA à GPA, alors que sa tranche commence à "
    "DCE : elle omet DCE, DAACT et ACT.",
    "C38 = C37+C35+C26+C24 n'inclut pas C19 : la ventilation de l'esquisse "
    "n'entre jamais au total, qui ne peut donc pas faire 100 %.",
    "D20 multiplie par D12 (taux optionnel) au lieu de D13 (base des "
    "honoraires) ; D19 porte 15 000 en dur quand les autres lignes sont des "
    "formules.",
]


def _borne(v, cote):
    """Une valeur du barème est une FOURCHETTE ; un marché porte un nombre.

    On ne moyenne pas les deux bornes en douce : le côté retenu est choisi par
    l'appelant et écrit dans le classeur. Moyenner produirait un chiffre que ni
    le bas ni le haut de la fourchette ne défend."""
    if isinstance(v, (list, tuple)):
        if not v:
            return 0.0
        return float(v[-1] if cote == "haut" else v[0])
    return float(v or 0.0)


def _missions_retenues(res):
    """Les missions du calcul qui portent réellement un montant."""
    out = []
    for m in res.get("missions") or []:
        if (m.get("part_retenue") or 0) <= 0:
            continue
        out.append(m)
    return out


def lignes_remplies(res, cote="haut"):
    """Le tableau, en données — la même matière que le classeur.

    Rendre les données AVANT le classeur permet de les contrôler sans ouvrir un
    fichier binaire, et à la page web d'afficher exactement ce qui sera
    téléchargé.
    """
    missions = _missions_retenues(res)
    base = _borne(res.get("total_meur"), cote) * 1e6
    out = []
    for L in LIGNES:
        e = {"code": L["code"], "nom": L["nom"], "montant": 0.0,
             "ventilation": 0.0, "par_mission": {}, "etat": "calcule", "dit": ""}
        if L.get("absent"):
            e["etat"] = "absent_du_bareme"
            e["dit"] = ("Aucune phase du barème ne couvre cet élément : il "
                        "reste à chiffrer hors de ce tableau.")
        elif L.get("groupe"):
            e["etat"] = "compris_dans"
            e["dit"] = "Compris dans %s — le barème ne partage pas ce groupe." % L["groupe"]
        elif L.get("mission"):
            m = next((x for x in missions if x["cle"] == L["mission"]), None)
            if not m:
                e["etat"] = "non_retenue"
                e["dit"] = "Mission non retenue dans ce calcul."
            else:
                e["montant"] = _borne(m.get("montant_meur"), cote) * 1e6
                e["par_mission"][m["cle"]] = e["montant"]
        else:
            ph = L["phase"]
            tot = 0.0
            for m in missions:
                # LA MISSION OPC A SA PROPRE LIGNE : la compter aussi dans les
                # phases la ferait payer deux fois.
                if m["cle"] == "opc":
                    continue
                p = (m.get("phases") or {}).get(ph) or {}
                if not p.get("retenue"):
                    continue
                v = _borne(p.get("montant_meur"), cote) * 1e6
                if v:
                    e["par_mission"][m["cle"]] = v
                    tot += v
            e["montant"] = tot
            if L["code"] in GROUPES:
                e["dit"] = ("Montant du groupe %s — le barème ne le partage pas."
                            % " + ".join(GROUPES[L["code"]]))
        out.append(e)
    # LA VENTILATION SE RAPPORTE AU TOTAL DU TABLEAU, PAS À LA BASE DU BARÈME.
    # Chaque montant du barème est arrondi au millier d'euros ; sur treize
    # missions et cinq phases, ces arrondis s'accumulent et la somme des lignes
    # s'écarte de la base contractuelle de quelques milliers d'euros. Rapportée
    # à la base, la colonne totalisait 100,08 % — un pourcentage qui ne fait pas
    # cent se remarque immédiatement dans une pièce de marché, et fait douter du
    # reste. Rapportée au total du tableau, elle fait exactement cent, et
    # L'ÉCART D'ARRONDI EST AFFICHÉ À PART plutôt que dissous dans les lignes.
    somme = sum(x["montant"] for x in out)
    for e in out:
        e["ventilation"] = (e["montant"] / somme) if somme else 0.0
    return out


def etat(res, cote="haut", operation="", reference=""):
    """Le bloc servi par l'API : de quoi afficher le tableau sans le fichier."""
    lg = lignes_remplies(res, cote)
    missions = _missions_retenues(res)
    total = sum(x["montant"] for x in lg)
    return {
        "ok": True,
        "version": VERSION,
        "operation": operation or "",
        "reference": reference or "",
        "cote": cote,
        "travaux_eur": _borne(res.get("travaux_meur"), cote) * 1e6,
        "base_honoraires_eur": _borne(res.get("total_meur"), cote) * 1e6,
        "taux_pct": _borne(res.get("taux_effectif_pct"), cote),
        "cotraitants": [{"cle": m["cle"], "nom": m["nom"]} for m in missions],
        "lignes": lg,
        "total_eur": total,
        # L'ÉCART D'ARRONDI EST PUBLIÉ, PAS ABSORBÉ. Forcer le total sur la base
        # obligerait à retoucher une ligne au hasard ; le taire laisserait un
        # lecteur attentif trouver seul une différence inexpliquée.
        "ecart_arrondi_eur": round(total - _borne(res.get("total_meur"), cote) * 1e6, 2),
        "ecart_arrondi_pct": round(
            (total / (_borne(res.get("total_meur"), cote) * 1e6) - 1) * 100, 3)
        if _borne(res.get("total_meur"), cote) else 0.0,
        "reserves_modele": RESERVES_MODELE,
        "limite": ("Le barème connaît cinq phases ; le modèle en demande "
                   "quatorze. Trois lignes sont des regroupements que la "
                   "source ne partage pas, et deux ne figurent pas au barème. "
                   "Rien n'y est réparti au prorata."),
    }


def classeur(res, cote="haut", operation="", reference="", date_maj=None):
    """Le classeur, construit sur la mise en page du modèle.

    On écrit des VALEURS et non les formules du modèle : celles-ci rendraient
    un total faux (voir l'en-tête). Le classeur reste lisible et vérifiable —
    chaque montant se recalcule depuis la ventilation et la base, toutes deux
    affichées.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    E = etat(res, cote, operation, reference)
    wb = Workbook()
    ws = wb.active
    ws.title = "Tableau de répartition"

    gras = Font(bold=True)
    petit = Font(size=9, color="5A5A5A")
    titre = Font(bold=True, size=13)
    fond_t = PatternFill("solid", fgColor="EFEFEA")
    fond_s = PatternFill("solid", fgColor="F7F6F3")
    fond_r = PatternFill("solid", fgColor="FFF9EF")
    bord = Border(bottom=Side(style="thin", color="D8D5CE"))
    droite = Alignment(horizontal="right")
    hautg = Alignment(vertical="top", wrap_text=True)

    ws["A2"] = "REF :";        ws["B2"] = E["reference"]
    ws["A3"] = "OPÉRATION :";  ws["B3"] = E["operation"]
    ws["A5"] = "Tableau de situation des honoraires de maîtrise d'œuvre"
    ws["A5"].font = titre
    ws["A7"] = ("Rempli depuis le barème du cabinet — aucune case n'est à "
                "saisir à la main.")
    ws["A7"].font = petit

    ws["A9"] = "Date de mise à jour du document :"
    ws["B9"] = (date_maj or _date.today()).isoformat()
    ws["A10"] = "Coût de construction HT marché (VRD incluse) :"
    ws["B10"] = round(E["travaux_eur"], 2)
    ws["A11"] = "Taux de rémunération :"
    ws["B11"] = round(E["taux_pct"] / 100.0, 5)
    ws["B11"].number_format = "0.00 %"
    ws["A12"] = "Base contractuelle des honoraires (€ HT) :"
    ws["B12"] = round(E["base_honoraires_eur"], 2)
    ws["A13"] = "Borne de la fourchette retenue :"
    ws["B13"] = ("haute" if cote == "haut" else "basse")
    for r in (10, 12):
        ws.cell(r, 2).number_format = '# ##0.00 €'
    for r in range(9, 14):
        ws.cell(r, 1).font = gras

    # ── En-tête du tableau : une paire de colonnes par cotraitant ───────────
    # AUTANT DE COLONNES QUE DE COTRAITANTS. Le modèle en fige sept ; le barème
    # peut en retenir treize. Se limiter à sept obligerait à fondre des
    # missions entre elles, ce qui effacerait précisément ce que ce tableau
    # doit montrer.
    lig_t = 16
    ws.cell(lig_t, 1, "Phase").font = gras
    ws.cell(lig_t, 2, "Intitulé").font = gras
    ws.cell(lig_t, 3, "Ventilation").font = gras
    ws.cell(lig_t, 4, "Montant € HT").font = gras
    col = 5
    for m in E["cotraitants"]:
        ws.cell(lig_t - 1, col, m["nom"]).font = gras
        ws.merge_cells(start_row=lig_t - 1, start_column=col,
                       end_row=lig_t - 1, end_column=col + 1)
        ws.cell(lig_t, col, "Ventilation").font = gras
        ws.cell(lig_t, col + 1, "Montant € HT").font = gras
        col += 2
    for c in range(1, col):
        ws.cell(lig_t, c).fill = fond_t
        ws.cell(lig_t, c).border = bord

    # ── Les lignes ─────────────────────────────────────────────────────────
    r = lig_t + 1
    for L in E["lignes"]:
        ws.cell(r, 1, L["code"]).font = gras
        ws.cell(r, 2, L["nom"])
        if L["etat"] == "calcule":
            ws.cell(r, 3, round(L["ventilation"], 5)).number_format = "0.00 %"
            ws.cell(r, 4, round(L["montant"], 2)).number_format = '# ##0.00 €'
            c = 5
            for m in E["cotraitants"]:
                v = L["par_mission"].get(m["cle"], 0.0)
                part = (v / L["montant"]) if L["montant"] else 0.0
                ws.cell(r, c, round(part, 5)).number_format = "0.00 %"
                ws.cell(r, c + 1, round(v, 2)).number_format = '# ##0.00 €'
                c += 2
        else:
            # NI ZÉRO NI PRORATA : la raison, en toutes lettres. Un zéro
            # laisserait croire que rien n'est dû pour cette phase.
            ws.cell(r, 3, L["dit"]).font = petit
            ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=col - 1)
            ws.cell(r, 3).alignment = hautg
        for c in range(1, col):
            ws.cell(r, c).border = bord
        r += 1

    ws.cell(r, 1, "TOTAL").font = gras
    ws.cell(r, 4, round(E["total_eur"], 2)).number_format = '# ##0.00 €'
    ws.cell(r, 4).font = gras
    c = 5
    for m in E["cotraitants"]:
        s = sum(x["par_mission"].get(m["cle"], 0.0) for x in E["lignes"])
        ws.cell(r, c + 1, round(s, 2)).number_format = '# ##0.00 €'
        ws.cell(r, c + 1).font = gras
        c += 2
    for cc in range(1, col):
        ws.cell(r, cc).fill = fond_s
    r += 2

    # ── Ce que le tableau ne dit pas ───────────────────────────────────────
    # L'écart d'arrondi, sous le total, là où on le cherche.
    if abs(E["ecart_arrondi_eur"]) >= 1:
        ws.cell(r - 1, 1, "Écart d'arrondi entre la somme des lignes et la base "
                          "contractuelle : %+.0f € (%+.3f %%). Chaque montant du "
                          "barème est arrondi au millier d'euros."
                          % (E["ecart_arrondi_eur"], E["ecart_arrondi_pct"])).font = petit
        ws.merge_cells(start_row=r - 1, start_column=1,
                       end_row=r - 1, end_column=max(6, col - 1))
        ws.cell(r - 1, 1).alignment = hautg
        r += 1

    ws.cell(r, 1, "Ce que ce tableau ne dit pas").font = gras
    r += 1
    for t in [E["limite"]] + ["Ligne « %s » : %s" % (x["code"], x["dit"])
                              for x in E["lignes"] if x["dit"]]:
        ws.cell(r, 1, t).alignment = hautg
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=max(6, col - 1))
        ws.cell(r, 1).fill = fond_r
        r += 1
    r += 1

    ws.cell(r, 1, "Réserves sur le modèle d'origine").font = gras
    r += 1
    ws.cell(r, 1, "Le classeur fourni porte des formules qui rendraient un "
                  "total faux quoi qu'on saisisse. Elles ne sont pas reprises ; "
                  "les montants ci-dessus sont des valeurs calculées.").font = petit
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=max(6, col - 1))
    ws.cell(r, 1).alignment = hautg
    r += 1
    for t in RESERVES_MODELE:
        ws.cell(r, 1, "— " + t).alignment = hautg
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=max(6, col - 1))
        r += 1

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 46
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 17
    for i in range(5, col):
        ws.column_dimensions[get_column_letter(i)].width = 15
    ws.freeze_panes = ws.cell(lig_t + 1, 3)
    return wb


def octets(res, cote="haut", operation="", reference="", date_maj=None):
    """Le classeur prêt à être servi."""
    tampon = io.BytesIO()
    classeur(res, cote, operation, reference, date_maj).save(tampon)
    tampon.seek(0)
    return tampon.getvalue()


def sante():
    return {"module": "repartition_honoraires", "version": VERSION,
            "lignes_modele": len(LIGNES),
            "regroupements": len(GROUPES),
            "absents_du_bareme": sum(1 for L in LIGNES if L.get("absent")),
            "reserves_modele": len(RESERVES_MODELE)}
